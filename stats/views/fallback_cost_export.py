from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .views_transportation_line_cost import build_transportation_line_fallback_analysis
from stats.excel_helpers import (
    make_workbook_styles,
    write_filter_row,
    write_summary_section,
    write_detail_header_row,
)
from stats.utils import parse_date, build_export_filename


@login_required
def export_fallback_cost_analysis_excel(request):
    date_from_str = request.GET.get("date_from", "")
    date_to_str = request.GET.get("date_to", "")
    vendor_name = request.GET.get("vendor_name", "").strip()
    item_no = request.GET.get("item_no", "").strip()

    date_from = parse_date(date_from_str)
    date_to = parse_date(date_to_str)

    fallback_analysis = build_transportation_line_fallback_analysis(
        date_from=date_from,
        date_to=date_to,
        vendor_name=vendor_name,
        item_no=item_no,
        include_rows=True,
    )

    wb = Workbook()
    styles = make_workbook_styles()
    header_fill, header_font, title_font, section_font, zebra_fill, border = styles

    # -----------------------------
    # Sheet 1: Summary
    # -----------------------------
    ws = wb.active
    ws.title = "Fallback Cost Analysis"

    summary_data = [
        ["Fallback Shipments",              fallback_analysis["cards"]["all"]],
        ["Air",                             fallback_analysis["cards"]["air"]],
        ["Sea",                             fallback_analysis["cards"]["sea"]],
        ["Road",                            fallback_analysis["cards"]["road"]],
        ["Courier",                         fallback_analysis["cards"]["courier"]],
        ["Other",                           fallback_analysis["cards"]["other"]],
        ["Fallback Goods Value (USD)",      float(fallback_analysis["summary"]["total_goods_usd"])],
        ["Fallback Transport Cost (USD)",   float(fallback_analysis["summary"]["total_transport_usd"])],
        ["Fallback Transport %",            float(fallback_analysis["summary"]["overall_percent"]) / 100],
    ]

    write_summary_section(ws, "Fallback Transportation Cost Analysis", summary_data, styles)
    write_filter_row(ws, date_from_str, date_to_str, vendor_name, item_no)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 20
    ws.freeze_panes = "A7"

    # -----------------------------
    # Sheet 2: Detailed analyzed rows
    # -----------------------------
    ws2 = wb.create_sheet(title="Fallback Rows")

    detail_headers = [
        "Import Code",
        "Vendor Name",
        "Shipping Method",
        "Declaration C Number",
        "Declaration Date",
        "Currency",
        "TRANSPORTATION Line Count",
        "Total Lines Amount",
        "Transportation Lines Amount",
        "Goods Excluding Transportation",
        "Total Lines USD",
        "Transportation USD",
        "Goods USD",
        "Transport %",
    ]

    write_detail_header_row(ws2, detail_headers, styles)

    for row_num, row in enumerate(fallback_analysis.get("rows", []), start=2):
        values = [
            row.get("import_code", ""),
            row.get("vendor_name", ""),
            row.get("shipping_method", ""),
            row.get("declaration_c_number", ""),
            row.get("declaration_date").strftime("%Y-%m-%d") if row.get("declaration_date") else "",
            row.get("currency_code", ""),
            int(row.get("transportation_line_count", 0) or 0),
            float(row.get("total_lines_amount", 0) or 0),
            float(row.get("transportation_lines_amount", 0) or 0),
            float(row.get("goods_amount", 0) or 0),
            float(row.get("total_lines_usd", 0) or 0),
            float(row.get("transport_usd", 0) or 0),
            float(row.get("goods_usd", 0) or 0),
            float(row.get("percent", 0) or 0) / 100,
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = zebra_fill
            if col_num in [8, 9, 10, 11, 12, 13]:
                cell.number_format = "0.00"
            if col_num == 14:
                cell.number_format = "0.00%"

    widths = {
        "A": 16, "B": 35, "C": 18, "D": 22, "E": 15,
        "F": 12, "G": 20, "H": 18, "I": 22, "J": 24,
        "K": 16, "L": 18, "M": 15, "N": 15,
    }
    for col_letter, width in widths.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{max(2, len(fallback_analysis.get('rows', [])) + 1)}"

    # -----------------------------
    # Sheet 3: Warnings
    # -----------------------------
    ws3 = wb.create_sheet(title="Warnings")
    ws3["A1"] = "Fallback Analysis Warnings"
    ws3["A1"].font = title_font

    ws3["A3"] = "Warning"
    ws3["A3"].fill = header_fill
    ws3["A3"].font = header_font
    from openpyxl.styles import Alignment
    ws3["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws3["A3"].border = border

    warnings_data = fallback_analysis.get("warnings", {})
    all_warnings = []
    for entry in warnings_data.get("skipped", []):
        code = entry.get("import_code") or "N/A"
        all_warnings.append(f"SKIPPED | {code} | {entry.get('reason', '')}")
    for entry in warnings_data.get("soft", []):
        code = entry.get("import_code") or "N/A"
        all_warnings.append(f"SOFT | {code} | {entry.get('reason', '')}")
    for entry in warnings_data.get("info", []):
        code = entry.get("import_code") or "N/A"
        all_warnings.append(f"INFO | {code} | {entry.get('reason', '')}")

    if all_warnings:
        for idx, warning in enumerate(all_warnings, start=4):
            cell = ws3.cell(row=idx, column=1, value=warning)
            cell.border = border
            if idx % 2 == 0:
                cell.fill = zebra_fill
    else:
        cell = ws3.cell(row=4, column=1, value="No warnings")
        cell.border = border

    ws3.column_dimensions["A"].width = 100

    filename = build_export_filename("fallback_cost_analysis", date_from_str, date_to_str, vendor_name, item_no)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
