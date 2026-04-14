from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from stats.cost_services import build_cost_analysis, build_unified_cost_analysis
from stats.excel_helpers import (
    make_workbook_styles,
    write_filter_row,
    write_summary_section,
    write_detail_header_row,
)
from stats.utils import parse_date, build_export_filename


@login_required
def export_cost_analysis_excel(request):
    date_from_str = request.GET.get("date_from", "")
    date_to_str = request.GET.get("date_to", "")
    vendor_name = request.GET.get("vendor_name", "").strip()
    item_no = request.GET.get("item_no", "").strip()

    date_from = parse_date(date_from_str)
    date_to = parse_date(date_to_str)

    analysis = build_cost_analysis(
        date_from=date_from,
        date_to=date_to,
        vendor_name=vendor_name,
        item_no=item_no,
        include_rows=True,
    )

    wb = Workbook()
    styles = make_workbook_styles()
    header_fill, header_font, title_font, section_font, zebra_fill, border = styles

    # -----------------------------
    # Sheet 1: Summary
    # -----------------------------
    ws = wb.active
    ws.title = "Cost Analysis"

    summary_data = [
        ["All Shipments",               analysis["cards"]["all"]],
        ["Air",                         analysis["cards"]["air"]],
        ["Sea",                         analysis["cards"]["sea"]],
        ["Road",                        analysis["cards"]["road"]],
        ["Courier",                     analysis["cards"]["courier"]],
        ["Other",                       analysis["cards"]["other"]],
        ["Total Goods Value (USD)",     float(analysis["summary"]["total_goods_usd"])],
        ["Total Transport Cost (USD)",  float(analysis["summary"]["total_transport_usd"])],
        ["Transport %",                 float(analysis["summary"]["overall_percent"]) / 100],
    ]

    write_summary_section(ws, "Transportation Cost Analysis", summary_data, styles)
    write_filter_row(ws, date_from_str, date_to_str, vendor_name, item_no)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 20
    ws.freeze_panes = "A7"

    # -----------------------------
    # Sheet 2: Detailed analyzed rows
    # -----------------------------
    ws2 = wb.create_sheet(title="Analyzed Rows")

    detail_headers = [
        "Import Code",
        "Vendor Name",
        "Shipping Method",
        "Declaration C Number",
        "Declaration Date",
        "Goods Amount",
        "Goods Currency",
        "Goods USD",
        "Transport Amount",
        "Transport Currency",
        "Transport USD",
        "Transport %",
    ]

    write_detail_header_row(ws2, detail_headers, styles)

    for row_num, row in enumerate(analysis.get("rows", []), start=2):
        values = [
            row.get("import_code", ""),
            row.get("vendor_name", ""),
            row.get("shipping_method", ""),
            row.get("declaration_c_number", ""),
            row.get("declaration_date").strftime("%Y-%m-%d") if row.get("declaration_date") else "",
            float(row.get("goods_amount", 0) or 0),
            row.get("goods_currency", ""),
            float(row.get("goods_usd", 0) or 0),
            float(row.get("transport_amount", 0) or 0),
            row.get("transport_currency", ""),
            float(row.get("transport_usd", 0) or 0),
            float(row.get("transport_percent", 0) or 0) / 100,
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = zebra_fill
            if col_num in [6, 8, 9, 11]:
                cell.number_format = "0.00"
            if col_num == 12:
                cell.number_format = "0.00%"

    widths = {
        "A": 16, "B": 35, "C": 18, "D": 20, "E": 15,
        "F": 15, "G": 14, "H": 15, "I": 18, "J": 18,
        "K": 15, "L": 15,
    }
    for col_letter, width in widths.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{max(2, len(analysis.get('rows', [])) + 1)}"

    filename = build_export_filename("cost_analysis", date_from_str, date_to_str, vendor_name, item_no)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_unified_cost_analysis_excel(request):
    date_from_str = request.GET.get("date_from", "")
    date_to_str = request.GET.get("date_to", "")
    vendor_name = request.GET.get("vendor_name", "").strip()
    item_no = request.GET.get("item_no", "").strip()

    date_from = parse_date(date_from_str)
    date_to = parse_date(date_to_str)

    analysis = build_unified_cost_analysis(
        date_from=date_from,
        date_to=date_to,
        vendor_name=vendor_name,
        item_no=item_no,
        include_rows=True,
    )

    wb = Workbook()
    styles = make_workbook_styles()
    header_fill, header_font, title_font, section_font, zebra_fill, border = styles

    # -----------------------------
    # Sheet 1: Summary
    # -----------------------------
    ws = wb.active
    ws.title = "Unified Cost Analysis"

    summary_data = [
        ["All Shipments",               analysis["cards"]["all"]],
        ["Air",                         analysis["cards"]["air"]],
        ["Sea",                         analysis["cards"]["sea"]],
        ["Road",                        analysis["cards"]["road"]],
        ["Courier",                     analysis["cards"]["courier"]],
        ["Other",                       analysis["cards"]["other"]],
        ["Total Goods Value (USD)",     float(analysis["summary"]["total_goods_usd"])],
        ["Total Transport Cost (USD)",  float(analysis["summary"]["total_transport_usd"])],
        ["Transport %",                 float(analysis["summary"]["overall_percent"]) / 100],
    ]

    write_summary_section(ws, "Unified Transportation Cost Analysis", summary_data, styles)
    write_filter_row(ws, date_from_str, date_to_str, vendor_name, item_no)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 20
    ws.freeze_panes = "A7"

    # -----------------------------
    # Sheet 2: Unified rows
    # -----------------------------
    ws2 = wb.create_sheet(title="Unified Rows")

    detail_headers = [
        "Import Code",
        "Vendor Name",
        "Shipping Method",
        "Declaration C Number",
        "Declaration Date",
        "Goods Amount",
        "Goods Currency",
        "Goods USD",
        "TRANSPORTATION Lines Amount",
        "Header Transport Amount",
        "Header Transport Currency",
        "Transport USD (Combined)",
        "Transport %",
    ]

    write_detail_header_row(ws2, detail_headers, styles)

    for row_num, row in enumerate(analysis.get("rows", []), start=2):
        values = [
            row.get("import_code", ""),
            row.get("vendor_name", ""),
            row.get("shipping_method", ""),
            row.get("declaration_c_number", ""),
            row.get("declaration_date").strftime("%Y-%m-%d") if row.get("declaration_date") else "",
            float(row.get("goods_amount", 0) or 0),
            row.get("goods_currency", ""),
            float(row.get("goods_usd", 0) or 0),
            float(row.get("transportation_lines_amount", 0) or 0),
            float(row.get("header_transport_amount", 0) or 0),
            row.get("header_transport_currency", ""),
            float(row.get("transport_usd", 0) or 0),
            float(row.get("transport_percent", 0) or 0) / 100,
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = zebra_fill
            # Numeric columns: Goods Amount, Goods USD,
            # TRANSPORTATION Lines Amount, Header Transport Amount, Transport USD
            if col_num in [6, 8, 9, 10, 12]:
                cell.number_format = "0.00"
            if col_num == 13:
                cell.number_format = "0.00%"

    widths = {
        "A": 16, "B": 35, "C": 18, "D": 22, "E": 15,
        "F": 15, "G": 14, "H": 15, "I": 28, "J": 22,
        "K": 22, "L": 22, "M": 15,
    }
    for col_letter, width in widths.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{max(2, len(analysis.get('rows', [])) + 1)}"

    filename = build_export_filename("unified_cost_analysis", date_from_str, date_to_str, vendor_name, item_no)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
