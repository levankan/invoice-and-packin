from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def make_workbook_styles():
    """
    Return the six shared style objects used across all export workbooks:
    (header_fill, header_font, title_font, section_font, zebra_fill, border)
    """
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)
    section_font = Font(size=12, bold=True)
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, header_font, title_font, section_font, zebra_fill, border


def write_filter_row(ws, date_from_str, date_to_str, vendor_name, item_no):
    """Write the Date From / Date To / Vendor / Item filter labels and values in row 3."""
    ws["A3"] = "Date From"
    ws["B3"] = date_from_str or "-"
    ws["C3"] = "Date To"
    ws["D3"] = date_to_str or "-"
    ws["E3"] = "Vendor Name"
    ws["F3"] = vendor_name or "-"
    ws["G3"] = "Item Number"
    ws["H3"] = item_no or "-"


def write_summary_section(ws, title, summary_rows, styles):
    """
    Write the standard summary block to ws:
      A1 = title (title_font)
      A5 = "Summary" (section_font)
      Row 7 = header row ["Metric", "Value"]
      Rows 8+ = data rows

    summary_rows: list of [label, value] pairs (int or float).
    Number format rules applied to column 2:
      - float where label ends with "(USD)"  → "0.00"
      - float where label ends with "%"      → "0.00%"
    """
    header_fill, header_font, title_font, section_font, zebra_fill, border = styles

    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A5"] = "Summary"
    ws["A5"].font = section_font

    start_row = 7
    for col_num, value in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=start_row, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index, row_data in enumerate(summary_rows, start_row + 1):
        label, value = row_data[0], row_data[1]
        for col_num, cell_value in enumerate([label, value], 1):
            cell = ws.cell(row=row_index, column=col_num, value=cell_value)
            cell.border = border
            if row_index % 2 == 0:
                cell.fill = zebra_fill
            if col_num == 2 and isinstance(cell_value, float):
                if label.endswith("(USD)"):
                    cell.number_format = "0.00"
                elif label.endswith("%"):
                    cell.number_format = "0.00%"


def write_detail_header_row(ws, headers, styles):
    """Write a styled header row in row 1 of a detail sheet."""
    header_fill, header_font, _, _, _, border = styles
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
