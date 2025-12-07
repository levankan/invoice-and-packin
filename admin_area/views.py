# admin_area/views.py
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse
from django.core.files.uploadedfile import UploadedFile
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
import math
import pandas as pd
from .forms import ItemsUploadForm, VendorsUploadForm
from .models import Item, Vendor
import csv
import os
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.shortcuts import render, redirect, get_object_or_404

from .models import Forwarder


def is_superuser(u):
    return u.is_superuser


# Helper: normalize possible NaN/blank strings to None (for DecimalFields etc.)
def none_if_nan(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("", "nan", "none", "null", "-"):
            return None
    return v


# ===== Admin Dashboard =====
@login_required
@user_passes_test(is_superuser)
def admin_dashboard(request):
    context = {"total_users": None, "recent_actions": []}
    return render(request, 'admin_area/dashboard.html', context)


# ===== User Management (stub) =====
@login_required
@user_passes_test(is_superuser)
def user_list(request):
    return HttpResponse("User Management page (coming soon)")


# ===== System Reports (stub) =====
@login_required
@user_passes_test(is_superuser)
def reports_overview(request):
    return HttpResponse("System Reports page (coming soon)")


# ===== ITEMS: List (search + pagination) =====
@login_required
@user_passes_test(is_superuser)
def items_view(request):
    q = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    items = Item.objects.all()
    if q:
        items = items.filter(
            Q(number__icontains=q) |
            Q(description__icontains=q) |
            Q(material__icontains=q) |
            Q(hs_code__icontains=q) |
            Q(item_category_code__icontains=q) |
            Q(parent_item_category__icontains=q) |
            Q(base_unit_of_measure__icontains=q) |
            Q(type__icontains=q)
        )

    items = items.order_by('-updated_at', 'number')

    paginator = Paginator(items, 25)
    page_obj = paginator.get_page(page_number)

    context = {"q": q, "page_obj": page_obj, "total": paginator.count}
    return render(request, 'admin_area/items.html', context)


# ===== ITEMS: Upload (Excel -> upsert by "No.") =====
@login_required
@user_passes_test(is_superuser)
def items_upload(request):
    if request.method == "POST":
        form = ItemsUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f: UploadedFile = form.cleaned_data['file']
            try:
                df = pd.read_excel(f, engine="openpyxl")
                df.columns = [c.strip().lower() for c in df.columns]

                colmap = {
                    'no.': 'number',
                    'description': 'description',
                    'parent item category': 'parent_item_category',
                    'base unit of measure': 'base_unit_of_measure',
                    'item category code': 'item_category_code',
                    'type': 'type',
                    'length': 'length',
                    'width': 'width',
                    'height': 'height',
                    'weight': 'weight',
                    'volumetric weight': 'volumetric_weight',
                    'material': 'material',
                    'hs code': 'hs_code',
                    'additional measurment': 'additional_measurement',
                }

                if 'no.' not in df.columns:
                    messages.error(request, "Missing required column: 'No.'")
                    return redirect('items_upload')

                used_cols = [c for c in df.columns if c in colmap]
                df = df[used_cols].rename(columns=colmap)

                df['number'] = df['number'].astype(str).str.strip()
                df = df[df['number'] != ""].copy()

                for num_col in ['length', 'width', 'height', 'weight', 'volumetric_weight']:
                    if num_col in df.columns:
                        df[num_col] = pd.to_numeric(df[num_col], errors='coerce')
                        df[num_col] = df[num_col].astype(object).where(~pd.isna(df[num_col]), None)

                numbers = df['number'].unique().tolist()
                existing_qs = Item.objects.filter(number__in=numbers)
                existing_map = {i.number: i for i in existing_qs}

                to_create, to_update = [], []
                for _, row in df.iterrows():
                    data = row.to_dict()
                    item_fields = {
                        'description': data.get('description') or "",
                        'parent_item_category': data.get('parent_item_category') or "",
                        'base_unit_of_measure': data.get('base_unit_of_measure') or "",
                        'item_category_code': data.get('item_category_code') or "",
                        'type': data.get('type') or "",
                        'length': none_if_nan(data.get('length')),
                        'width': none_if_nan(data.get('width')),
                        'height': none_if_nan(data.get('height')),
                        'weight': none_if_nan(data.get('weight')),
                        'volumetric_weight': none_if_nan(data.get('volumetric_weight')),
                        'material': data.get('material') or "",
                        'hs_code': ("" if none_if_nan(data.get('hs_code')) is None else str(data.get('hs_code')).strip()),
                        'additional_measurement': ("" if none_if_nan(data.get('additional_measurement')) is None else str(data.get('additional_measurement')).strip()),
                    }
                    number = data['number']
                    if number in existing_map:
                        obj = existing_map[number]
                        for k, v in item_fields.items():
                            setattr(obj, k, v)
                        to_update.append(obj)
                    else:
                        to_create.append(Item(number=number, **item_fields))

                with transaction.atomic():
                    created = Item.objects.bulk_create(to_create, ignore_conflicts=True)
                    if to_update:
                        Item.objects.bulk_update(
                            to_update,
                            fields=['description', 'parent_item_category', 'base_unit_of_measure',
                                    'item_category_code', 'type', 'length', 'width', 'height',
                                    'weight', 'volumetric_weight', 'material', 'hs_code', 'additional_measurement'],
                            batch_size=500,
                        )

                messages.success(request, f"Items imported successfully. Created: {len(created)}, Updated: {len(to_update)}.")
                return redirect('items_view')

            except Exception as e:
                messages.error(request, f"Import failed: {e}")
                return redirect('items_upload')
    else:
        form = ItemsUploadForm()

    return render(request, 'admin_area/items_upload.html', {'form': form})


# ===== VENDORS: List (search + pagination) =====
@login_required
@user_passes_test(is_superuser)
def vendors_view(request):
    q = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    vendors = Vendor.objects.all()
    if q:
        vendors = vendors.filter(
            Q(number__icontains=q) |
            Q(name__icontains=q) |
            Q(vat_registration_no__icontains=q)
        )

    vendors = vendors.order_by('name', 'number')

    paginator = Paginator(vendors, 25)
    page_obj = paginator.get_page(page_number)

    context = {"q": q, "page_obj": page_obj, "total": paginator.count}
    return render(request, 'admin_area/vendors.html', context)


# ===== VENDORS: Upload (Excel -> upsert by "No.") =====
@login_required
@user_passes_test(is_superuser)
def vendors_upload(request):
    if request.method == "POST":
        form = VendorsUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data['file']
            try:
                df = pd.read_excel(f, engine="openpyxl")
                df.columns = [c.strip().lower() for c in df.columns]

                required = ['no.', 'name']
                for col in required:
                    if col not in df.columns:
                        messages.error(request, f"Missing required column: {col}")
                        return redirect('vendors_upload')

                df = df.rename(columns={
                    'no.': 'number',
                    'name': 'name',
                    'vat registration no.': 'vat_registration_no',
                })

                df['number'] = df['number'].astype(str).str.strip()
                df['name'] = df['name'].astype(str).str.strip()
                df = df[(df['number'] != "") & (df['name'] != "")].copy()

                existing = {v.number: v for v in Vendor.objects.filter(number__in=df['number'])}
                to_create, to_update = [], []

                for _, row in df.iterrows():
                    data = row.to_dict()
                    fields = {
                        'name': data.get('name', ''),
                        'vat_registration_no': str(data.get('vat_registration_no') or '').strip(),
                    }
                    num = data['number']
                    if num in existing:
                        obj = existing[num]
                        for k, v in fields.items():
                            setattr(obj, k, v)
                        to_update.append(obj)
                    else:
                        to_create.append(Vendor(number=num, **fields))

                with transaction.atomic():
                    created = Vendor.objects.bulk_create(to_create, ignore_conflicts=True)
                    if to_update:
                        Vendor.objects.bulk_update(to_update, ['name', 'vat_registration_no'], batch_size=500)

                messages.success(request, f"Vendors imported successfully. Created: {len(to_create)}, Updated: {len(to_update)}.")
                return redirect('vendors_view')

            except Exception as e:
                messages.error(request, f"Import failed: {e}")
                return redirect('vendors_upload')
    else:
        form = VendorsUploadForm()

    return render(request, 'admin_area/vendors_upload.html', {'form': form})








# Only superuser can access admin area
def is_superuser(user):
    return user.is_superuser


@login_required
@user_passes_test(is_superuser)
def forwarders_view(request):
    """
    List forwarders. Upload is handled by forwarders_upload().
    """
    forwarders = Forwarder.objects.all().order_by("name")
    return render(
        request,
        "admin_area/forwarders.html",
        {"forwarders": forwarders},
    )



@login_required
@user_passes_test(is_superuser)
def forwarders_export(request):
    """
    Download all forwarders as an Excel file (.xlsx).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Forwarders"

    # Header row
    ws.append(["name", "Legal Name", "VAT Registration No."])

    # Data rows
    for f in Forwarder.objects.all().order_by("name"):
        ws.append([
            f.name,
            f.legal_name or "",
            f.vat_registration_no or "",
        ])

    # Prepare HTTP response
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = 'attachment; filename="forwarders.xlsx"'
    wb.save(response)
    return response



@login_required
@user_passes_test(is_superuser)
def forwarder_delete(request, pk):
    """
    Delete a forwarder (POST only).
    """
    if request.method == "POST":
        fwd = get_object_or_404(Forwarder, pk=pk)
        fwd.delete()
    return redirect("forwarders_view")


@login_required
@user_passes_test(is_superuser)
def forwarder_edit(request, pk):
    """
    Simple edit page for one forwarder.
    """
    fwd = get_object_or_404(Forwarder, pk=pk)

    if request.method == "POST":
        fwd.name = (request.POST.get("name") or "").strip()
        fwd.legal_name = (request.POST.get("legal_name") or "").strip() or None
        fwd.vat_registration_no = (request.POST.get("vat_registration_no") or "").strip() or None
        fwd.save()
        return redirect("forwarders_view")

    return render(request, "admin_area/forwarder_edit.html", {"forwarder": fwd})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def forwarders_upload(request):
    """
    Upload Forwarders from an Excel file.

    Expected columns in the first row:
    name | legal_name | vat_registration_no
    (header row can be anything, we just use column positions)
    """
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("forwarders_view")

    upload_file = request.FILES.get("file")
    if not upload_file:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("forwarders_view")

    try:
        wb = openpyxl.load_workbook(upload_file)
    except Exception:
        messages.error(request, "Could not read the file. Make sure it is a valid .xlsx Excel file.")
        return redirect("forwarders_view")

    sheet = wb.active
    created = 0
    updated = 0

    # We expect: first column = name, second = legal_name, third = vat_registration_no
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name, legal_name, vat_no = row[:3]

        if not name:
            continue  # skip rows without name

        name = str(name).strip()
        legal_name = (str(legal_name).strip() if legal_name else "")
        vat_no = (str(vat_no).strip() if vat_no else "")

        fwd, was_created = Forwarder.objects.update_or_create(
            name=name,
            defaults={
                "legal_name": legal_name or None,
                "vat_registration_no": vat_no or None,
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    messages.success(
        request,
        f"Forwarders upload completed. Created: {created}, Updated: {updated}."
    )
    return redirect("forwarders_view")





@login_required
@user_passes_test(lambda u: u.is_superuser)
def forwarder_create(request):
    """
    Create forwarder manually.
    """
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        legal_name = (request.POST.get("legal_name") or "").strip()
        vat_no = (request.POST.get("vat_registration_no") or "").strip()

        if not name:
            messages.error(request, "Forwarder name is required.")
        else:
            Forwarder.objects.create(
                name=name,
                legal_name=legal_name or None,
                vat_registration_no=vat_no or None,
            )
            messages.success(request, "Forwarder created successfully.")
            return redirect("forwarders_view")

    return render(request, "admin_area/forwarder_create.html")
