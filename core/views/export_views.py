from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from decimal import Decimal
import openpyxl
import pandas as pd

from core.models import Export, LineItem, Pallet
from core.forms import ExportForm
from .generate_doc_view import EXPECTED_HEADERS, EXPECTED_PALLET_HEADERS


# ==========================
# List Exports
# ==========================
@login_required
def exports_view(request):
    """List and search exports"""
    q = request.GET.get("q", "")
    exports = Export.objects.all().order_by("-created_at")
    if q:
        exports = exports.filter(
            Q(invoice_number__icontains=q)
            | Q(project_no__icontains=q)
            | Q(export_number__icontains=q)
        )
    return render(request, "core/exports.html", {"exports": exports})


# ==========================
# Edit Export
# ==========================
@login_required
def edit_export(request, export_id):
    """Edit a single export (with option to re-upload Excel data)."""
    export = get_object_or_404(Export, id=export_id)

    if request.method == "POST":
        form = ExportForm(request.POST, instance=export)
        file = request.FILES.get("excel_file")

        if form.is_valid():
            export = form.save(commit=False)

            # ✅ Save declaration fields manually
            export.declaration_c_number = request.POST.get("declaration_c_number")
            export.declaration_a_number = request.POST.get("declaration_a_number")
            export.declaration_register_date = (
                request.POST.get("declaration_register_date") or None
            )

            # ✅ Save who edited it (only if empty)
            if not export.created_by:
                export.created_by = request.user

            export.save()

            # Handle uploaded Excel
            if file:
                try:
                    xls = pd.ExcelFile(file)
                except Exception:
                    messages.error(
                        request,
                        "⚠ Could not read Excel file. Make sure it's .xlsx or .xls",
                    )
                    return redirect("edit_export", export_id=export.id)

                # Sheet1 → Line Items
                if "Sheet1" in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name="Sheet1")

                    if list(df.columns) != EXPECTED_HEADERS:
                        messages.error(
                            request, "⚠ Sheet1 headers do not match expected format."
                        )
                        return redirect("edit_export", export_id=export.id)

                    export.items.all().delete()

                    for _, row in df.iterrows():
                        LineItem.objects.create(
                            export=export,
                            serial_lot_number=row["Serial/Lot Number"],
                            document_number=row["Document Number"],
                            item_number=row["Item Number"],
                            cross_reference=row["Cross Reference #"],
                            qty=row["QTY"],
                            unit_of_measure=row["Unit of Measure"],
                            box_number=row["Box #"],
                            commercial_invoice_number=row["Commercial Invoice #"],
                            posting_date=row["Posting Date"],
                            shipment_number=row["Shipment #"],
                            description=row["Description"],
                            carbon_qty=row["Carbon QTY"],
                            carbon_lot=row["Carbon LOT"],
                            customer_po=row["Customer PO"],
                            po_line=row["PO Line"],
                            sales_order=row["Sales Order"],
                            sales_order_line=row["Sales Order Line"],
                            pallet_number=row["Pallet #"],
                            price=row["Price"],
                            lu=row["LU"],
                        )

                # Sheet2 → Pallets
                if "Sheet2" in xls.sheet_names:
                    df2 = pd.read_excel(xls, sheet_name="Sheet2")

                    if list(df2.columns) == EXPECTED_PALLET_HEADERS:
                        export.pallets.all().delete()
                        for _, row in df2.iterrows():
                            Pallet.objects.create(
                                export=export,
                                pallet_number=row["Pallet #"],
                                length_cm=Decimal(str(row["Lenght (Cm)"]))
                                if row["Lenght (Cm)"] is not None
                                else None,
                                width_cm=Decimal(str(row["Width (Cm)"]))
                                if row["Width (Cm)"] is not None
                                else None,
                                height_cm=Decimal(str(row["Height (Cm)"]))
                                if row["Height (Cm)"] is not None
                                else None,
                                gross_weight_kg=Decimal(
                                    str(row["Gross Weight (Kg)"])
                                )
                                if row["Gross Weight (Kg)"] is not None
                                else None,
                            )

            messages.success(request, "✅ Export updated successfully")
            return redirect("exports_view")

    else:
        form = ExportForm(instance=export)

    return render(request, "core/edit_export.html", {"form": form, "export": export})


# ==========================
# Delete Export
# ==========================
@login_required
def delete_export(request, export_id):
    if not request.user.is_superuser:
        return HttpResponseForbidden("You do not have permission to delete exports.")

    export = get_object_or_404(Export, id=export_id)

    if request.method == "POST":
        export.delete()
        return redirect("exports_view")

    return render(request, "core/delete_export.html", {"export": export})


# ==========================
# Download Excel Template
# ==========================
@login_required
def download_export_template(request):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    headers_items = [
        "Serial/Lot Number",
        "Document Number",
        "Item Number",
        "Cross Reference #",
        "QTY",
        "Unit of Measure",
        "Box #",
        "Commercial Invoice #",
        "Posting Date",
        "Shipment #",
        "Description",
        "Carbon QTY",
        "Carbon LOT",
        "Customer PO",
        "PO Line",
        "Sales Order",
        "Sales Order Line",
        "Pallet #",
        "Price",
        "LU",
    ]

    for col_num, header in enumerate(headers_items, 1):
        ws1.cell(row=1, column=col_num, value=header)

    ws2 = wb.create_sheet(title="Sheet2")
    headers_pallets = ["Pallet #", "Lenght (Cm)", "Width (Cm)", "Height (Cm)", "Gross Weight (Kg)"]

    for col_num, header in enumerate(headers_pallets, 1):
        ws2.cell(row=1, column=col_num, value=header)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Export_Template.xlsx"'
    wb.save(response)
    return response


# ==========================
# Export Database Excel
# ==========================
@login_required
def export_database_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exports Full"

    headers = [
        "Serial/Lot Number",
        "Document Number",
        "Item Number",
        "Cross Reference #",
        "QTY",
        "Unit of Measure",
        "Box #",
        "Commercial Invoice #",
        "Posting Date",
        "Shipment #",
        "Description",
        "Carbon QTY",
        "Carbon LOT",
        "Customer PO",
        "PO Line",
        "Sales Order",
        "Sales Order Line",
        "Pallet #",
        "Price",
        "LU",
        "Invoice No",
        "Packing List No",
        "Export No",
        "Project No",
        "Created At",
        "Created By",
        "Declaration C No",
        "Declaration A No",
        "Declaration Register Date",
        "Seller",
        "Sold To",
        "Shipped To",
        "Export ID",
    ]
    ws.append(headers)

    for exp in Export.objects.all().prefetch_related("items").order_by("id"):
        for item in exp.items.all():
            ws.append([
                item.serial_lot_number or "",
                item.document_number or "",
                item.item_number or "",
                item.cross_reference or "",
                item.qty or 0,
                item.unit_of_measure or "",
                item.box_number or "",
                item.commercial_invoice_number or "",
                item.posting_date.strftime("%Y-%m-%d") if item.posting_date else "",
                item.shipment_number or "",
                item.description or "",
                item.carbon_qty or "",
                item.carbon_lot or "",
                item.customer_po or "",
                item.po_line or "",
                item.sales_order or "",
                item.sales_order_line or "",
                item.pallet_number or "",
                item.price or 0,
                item.lu or "",
                exp.invoice_number or "",
                exp.packing_list_number or "",
                exp.export_number or "",
                exp.project_no or "",
                exp.created_at.strftime("%Y-%m-%d"),
                exp.created_by.username if exp.created_by else "",
                exp.declaration_c_number or "",
                exp.declaration_a_number or "",
                exp.declaration_register_date.strftime("%Y-%m-%d") if exp.declaration_register_date else "",
                exp.seller or "",
                exp.sold_to or "",
                exp.shipped_to or "",
                exp.id,
            ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Exports_Database.xlsx"'
    wb.save(response)
    return response


# ==========================
# Export Detail
# ==========================
@login_required
def export_detail(request, export_id):
    export = get_object_or_404(Export, id=export_id)
    return render(request, "core/export_success.html", {"export": export})
