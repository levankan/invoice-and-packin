# imports/views/lines.py
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect

from ..models import Import, ImportLine
from ..permissions import has_imports_access


@login_required
@user_passes_test(has_imports_access)
def upload_import_lines(request, pk):
    """
    Upload Excel lines for a given Import (pk) (edit mode).
    """
    imp = get_object_or_404(Import, pk=pk)

    if request.method == "POST" and request.FILES.get("file"):
        upload = request.FILES["file"]
        filename = upload.name.lower()

        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
            return redirect("imports_edit", pk=imp.pk)

        try:
            wb = openpyxl.load_workbook(upload, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Could not read the Excel file. Please check the format.")
            return redirect("imports_edit", pk=imp.pk)

        # ---- Build header map ----
        header_map = {}
        header_row = 1
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=header_row, column=col).value
            if raw is None:
                continue
            name = str(raw).strip()
            header_map[name] = col

        required_cols = [
            "Document No.",
            "Line No.",
            "No.",
            "Description",
            "Quantity",
            "Unit of Measure",
            "Direct Unit Cost Excl. VAT",
            "Line Amount Excl. VAT",
            "Expected Receipt Date",
            "Delivery Date",
        ]
        missing = [c for c in required_cols if c not in header_map]
        if missing:
            messages.error(
                request,
                "Missing required columns in Excel: " + ", ".join(missing)
            )
            return redirect("imports_edit", pk=imp.pk)

        # Helpers
        def _get(row, col_name):
            col_idx = header_map.get(col_name)
            if not col_idx:
                return None
            return ws.cell(row=row, column=col_idx).value

        def _clean_str(v):
            if v is None:
                return None
            s = str(v).strip()
            return s or None

        def _clean_decimal(v):
            if v is None or v == "":
                return None
            try:
                return Decimal(str(v).replace(",", ""))
            except (InvalidOperation, TypeError):
                return None

        def _clean_date(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip()
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        created_count = 0

        # ---- Read data rows ----
        for row in range(header_row + 1, ws.max_row + 1):
            doc_no = _clean_str(_get(row, "Document No."))
            line_no = _clean_str(_get(row, "Line No."))
            item_no = _clean_str(_get(row, "No."))
            desc = _clean_str(_get(row, "Description"))
            qty = _clean_decimal(_get(row, "Quantity"))
            uom = _clean_str(_get(row, "Unit of Measure"))
            unit_cost = _clean_decimal(_get(row, "Direct Unit Cost Excl. VAT"))
            line_amount = _clean_decimal(_get(row, "Line Amount Excl. VAT"))
            exp_date = _clean_date(_get(row, "Expected Receipt Date"))
            deliv_date = _clean_date(_get(row, "Delivery Date"))

            if not (doc_no or item_no or desc or qty):
                continue

            ImportLine.objects.create(
                import_header=imp,
                document_no=doc_no,
                line_no=line_no,
                item_no=item_no,
                description=desc,
                quantity=qty,
                unit_of_measure=uom,
                unit_cost=unit_cost,
                line_amount=line_amount,
                expected_receipt_date=exp_date,
                delivery_date=deliv_date,
            )
            created_count += 1

        messages.success(
            request,
            f"Successfully imported {created_count} lines for {imp.import_code}."
        )
        return redirect("imports_edit", pk=imp.pk)

    return redirect("imports_edit", pk=imp.pk)
