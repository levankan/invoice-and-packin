# imports/views/payments_excel.py
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

from ..models import Import
from ..permissions import has_imports_access


@login_required
def export_payments_excel(request):
    """
    Export payments/charges data to Excel.
    Columns: import_code + all charges area fields.
    """
    q = (request.GET.get("q") or "").strip()
    status_f = request.GET.getlist("status")
    method_f = request.GET.getlist("method")

    qs = (
        Import.objects
        .select_related(
            "transport_forwarder",
            "brokerage_forwarder",
            "internal_delivery_forwarder",
            "other1_forwarder",
            "other2_forwarder",
        )
        .order_by("-created_at")
    )

    # (Optional) keep same dashboard filters
    if q:
        qs = qs.filter(
            Q(import_code__icontains=q)
            | Q(vendor_name__icontains=q)
            | Q(tracking_no__icontains=q)
            | Q(vendor_reference__icontains=q)
            | Q(forwarder_reference__icontains=q)
        ).distinct()

    if status_f:
        qs = qs.filter(shipment_status__in=status_f)

    if method_f:
        qs = qs.filter(shipping_method__in=method_f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = [
        "Import Code",

        "Transport Forwarder",
        "Transport Invoice No",
        "Transport Price",
        "Transport Currency",
        "Transport Payment Date",

        "Brokerage Forwarder",
        "Brokerage Invoice No",
        "Brokerage Price",
        "Brokerage Currency",
        "Brokerage Payment Date",

        "Internal Delivery Forwarder",
        "Internal Delivery Invoice No",
        "Internal Delivery Price",
        "Internal Delivery Currency",
        "Internal Delivery Payment Date",

        "Other1 Forwarder",
        "Other1 Invoice No",
        "Other1 Price",
        "Other1 Currency",
        "Other1 Payment Date",

        "Other2 Forwarder",
        "Other2 Invoice No",
        "Other2 Price",
        "Other2 Currency",
        "Other2 Payment Date",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for imp in qs:
        row = [
            imp.import_code,

            imp.transport_forwarder.name if imp.transport_forwarder else "",
            imp.transport_invoice_no or "",
            float(imp.transport_price) if imp.transport_price is not None else "",
            imp.transport_currency or "",
            imp.transport_payment_date.isoformat() if imp.transport_payment_date else "",

            imp.brokerage_forwarder.name if imp.brokerage_forwarder else "",
            imp.brokerage_invoice_no or "",
            float(imp.brokerage_price) if imp.brokerage_price is not None else "",
            imp.brokerage_currency or "",
            imp.brokerage_payment_date.isoformat() if imp.brokerage_payment_date else "",

            imp.internal_delivery_forwarder.name if imp.internal_delivery_forwarder else "",
            imp.internal_delivery_invoice_no or "",
            float(imp.internal_delivery_price) if imp.internal_delivery_price is not None else "",
            imp.internal_delivery_currency or "",
            imp.internal_delivery_payment_date.isoformat() if imp.internal_delivery_payment_date else "",

            imp.other1_forwarder.name if imp.other1_forwarder else "",
            imp.other1_invoice_no or "",
            float(imp.other1_price) if imp.other1_price is not None else "",
            imp.other1_currency or "",
            imp.other1_payment_date.isoformat() if imp.other1_payment_date else "",

            imp.other2_forwarder.name if imp.other2_forwarder else "",
            imp.other2_invoice_no or "",
            float(imp.other2_price) if imp.other2_price is not None else "",
            imp.other2_currency or "",
            imp.other2_payment_date.isoformat() if imp.other2_payment_date else "",
        ]
        ws.append(row)

        # row style
        r = ws.max_row
        fill_color = "FFFFFF" if r % 2 == 0 else "F7F7F7"
        for c in ws[r]:
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            c.border = thin_border

    # auto-fit
    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="imports_payments.xlsx"'
    wb.save(response)
    return response
