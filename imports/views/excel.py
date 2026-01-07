# imports/views/excel.py

from collections import defaultdict
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from imports.models import Import, ImportLine, ImportPackage
from admin_area.models import Item


# --------------------------
# Helpers (styles)
# --------------------------
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
THIN_BORDER = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)


@login_required
def export_imports_excel(request):
    """
    Export Imports list (same filters as dashboard) to Excel.
    """
    q = (request.GET.get("q") or "").strip()
    status_f = request.GET.getlist("status")
    method_f = request.GET.getlist("method")

    qs = Import.objects.select_related("forwarder").order_by("-created_at")

    if q:
        qs = qs.filter(
            Q(import_code__icontains=q)
            | Q(vendor_name__icontains=q)
            | Q(tracking_no__icontains=q)
            | Q(vendor_reference__icontains=q)
            | Q(forwarder_reference__icontains=q)
            | Q(lines__item_no__icontains=q)
            | Q(lines__document_no__icontains=q)
        ).distinct()

    if status_f:
        qs = qs.filter(shipment_status__in=status_f)

    if method_f:
        qs = qs.filter(shipping_method__in=method_f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Imports"

    headers = [
        "DB ID",
        "Import Code",
        "Vendor",
        "Exporter Country",
        "Incoterms",
        "Currency",
        "Goods Price",
        "Shipping Method",
        "Shipment Status",
        "Forwarder",
        "Vendor Reference",
        "Forwarder Reference",
        "Tracking Number",
        "Pickup Address",
        "Is Danger",
        "Is Stackable",
        "Notes",
        "Declaration C Number",
        "Declaration A Number",
        "Declaration Date",
        "ATC Receipt Date",
        "Created At",
        "Total Gross Weight (kg)",
        "Total Volumetric Weight (kg)",
        "Forwarder VRS",  # ✅ NEW LAST COLUMN
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for imp in qs:
        # ✅ Forwarder VRS = Register Date (created_at) - Declaration Date (days)
        forwarder_vrs = ""
        if imp.created_at and imp.declaration_date:
            forwarder_vrs = (imp.created_at.date() - imp.declaration_date).days

        ws.append([
            imp.pk,
            imp.import_code,
            imp.vendor_name or "",
            str(imp.exporter_country) if imp.exporter_country else "",
            imp.incoterms or "",
            imp.currency_code or "",
            float(imp.goods_price) if imp.goods_price is not None else "",
            imp.shipping_method or "",
            imp.shipment_status or "",
            imp.forwarder.name if getattr(imp, "forwarder", None) else "",
            imp.vendor_reference or "",
            imp.forwarder_reference or "",
            imp.tracking_no or "",
            (imp.pickup_address or "").replace("\r\n", " ").replace("\n", " "),
            "YES" if imp.is_danger else "NO",
            "YES" if imp.is_stackable else "NO",
            (imp.notes or "").replace("\r\n", " ").replace("\n", " "),
            imp.declaration_c_number or "",
            imp.declaration_a_number or "",
            imp.declaration_date.isoformat() if imp.declaration_date else "",
            imp.expected_receipt_date.isoformat() if imp.expected_receipt_date else "",
            imp.created_at.strftime("%Y-%m-%d %H:%M") if imp.created_at else "",
            float(imp.total_gross_weight_kg) if imp.total_gross_weight_kg is not None else "",
            float(imp.total_volumetric_weight_kg) if imp.total_volumetric_weight_kg is not None else "",
            forwarder_vrs,  # ✅ NEW VALUE AT END
        ])

        row_idx = ws.max_row
        fill_color = "FFFFFF" if row_idx % 2 == 0 else "F7F7F7"
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = THIN_BORDER

    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 2, 60)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="imports_export.xlsx"'
    wb.save(response)
    return response



@login_required
def export_import_lines_excel(request):
    """
    Export all import lines to Excel including:
    - Vendor VRS = Register Date - Delivery Date
    - Forwarder VRS = Register Date - Declaration Date ✅
    """
    qs = (
        ImportLine.objects
        .select_related("import_header")
        .order_by("-import_header__created_at", "document_no", "line_no")
    )

    item_numbers = {line.item_no for line in qs if line.item_no}
    items = Item.objects.filter(number__in=item_numbers)
    item_map = {i.number: i for i in items}

    total_qty_by_import = defaultdict(Decimal)
    for line in qs:
        if line.import_header_id and line.quantity is not None:
            total_qty_by_import[line.import_header_id] += line.quantity

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Lines"

    headers = [
        "Line ID",
        "Import ID",
        "Import Code",
        "Vendor",
        "Document No.",
        "Line No.",
        "Item No.",
        "Description",
        "Quantity",
        "Unit of Measure",
        "Unit Cost",
        "Line Amount",
        "Goods Currency",
        "Expected Receipt Date",
        "Delivery Date",
        "ATC Receipt Date",
        "Created At",
        "Shipment Status",
        "Tracking Number",
        "Vendor Reference",
        "Forwarder Reference",
        "Declaration C Number",
        "Declaration A Number",
        "Declaration Date",
        "Incoterms",
        "Total Import GW (kg)",
        "Total Import VW (kg)",
        "Line Gross Weight (kg)",
        "Line Volumetric Weight (kg)",
        "Vendor VRS",
        "Forwarder VRS",  # ✅ NEW LAST COLUMN
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for line in qs:
        imp = line.import_header
        qty = line.quantity or Decimal("0")

        total_gw_import = imp.total_gross_weight_kg if imp else None
        total_vw_import = imp.total_volumetric_weight_kg if imp else None
        total_qty = total_qty_by_import.get(imp.id if imp else None, Decimal("0"))

        line_gw = None
        line_vw = None

        item = item_map.get(line.item_no) if line.item_no else None

        if qty and imp:
            if item and item.weight is not None:
                line_gw = item.weight * qty
            elif total_gw_import is not None and total_qty:
                line_gw = (total_gw_import * qty) / total_qty

            if item and item.volumetric_weight is not None:
                line_vw = item.volumetric_weight * qty
            elif total_vw_import is not None and total_qty:
                line_vw = (total_vw_import * qty) / total_qty

        register_dt = imp.created_at if imp else None

        vendor_vrs = ""
        if register_dt and line.delivery_date:
            vendor_vrs = (register_dt.date() - line.delivery_date).days

        forwarder_vrs = ""
        if register_dt and imp and imp.declaration_date:
            forwarder_vrs = (register_dt.date() - imp.declaration_date).days

        ws.append([
            line.pk,
            imp.pk if imp else "",
            imp.import_code if imp else "",
            imp.vendor_name if imp else "",
            line.document_no or "",
            line.line_no or "",
            line.item_no or "",
            line.description or "",
            float(line.quantity) if line.quantity is not None else "",
            line.unit_of_measure or "",
            float(line.unit_cost) if line.unit_cost is not None else "",
            float(line.line_amount) if line.line_amount is not None else "",
            (imp.currency_code or "") if imp else "",
            line.expected_receipt_date.isoformat() if line.expected_receipt_date else "",
            line.delivery_date.isoformat() if line.delivery_date else "",
            imp.expected_receipt_date.isoformat() if (imp and imp.expected_receipt_date) else "",
            imp.created_at.strftime("%Y-%m-%d") if (imp and imp.created_at) else "",
            (imp.shipment_status or "") if imp else "",
            imp.tracking_no if (imp and imp.tracking_no) else "",
            (imp.vendor_reference or "") if imp else "",
            (imp.forwarder_reference or "") if imp else "",
            (imp.declaration_c_number or "") if imp else "",
            (imp.declaration_a_number or "") if imp else "",
            imp.declaration_date.isoformat() if (imp and imp.declaration_date) else "",
            (imp.incoterms or "") if imp else "",
            float(total_gw_import) if total_gw_import is not None else "",
            float(total_vw_import) if total_vw_import is not None else "",
            float(line_gw) if line_gw is not None else "",
            float(line_vw) if line_vw is not None else "",
            vendor_vrs,
            forwarder_vrs,
        ])

        row_idx = ws.max_row
        fill_color = "FFFFFF" if row_idx % 2 == 0 else "F7F7F7"
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = THIN_BORDER

    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 2, 60)

    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="import_lines.xlsx"'
    wb.save(response)
    return response





@login_required
def export_single_import_excel(request, import_id):
    """
    Export single import:
    - Sheet1: header + lines
    - Sheet2: packages
    Includes Vendor VRS + Forwarder VRS ✅
    """
    imp = get_object_or_404(
        Import.objects.select_related("forwarder").prefetch_related("lines"),
        pk=import_id,
    )
    lines = imp.lines.all().order_by("document_no", "line_no")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Import {imp.import_code}"

    # ✅ Forwarder VRS for PART 1 (Register Date - Declaration Date)
    forwarder_vrs_part1 = ""
    if imp.created_at and imp.declaration_date:
        forwarder_vrs_part1 = (imp.created_at.date() - imp.declaration_date).days

    # --------------------------
    # Sheet1 - Part 1 header
    # --------------------------
    header_row_1 = [
        "Import ID", "Import Code", "Vendor", "Exporter Country", "Incoterms", "Currency",
        "Goods Price", "Shipping Method", "Shipment Status", "Tracking Number",
        "Vendor Reference", "Forwarder Reference", "Declaration C Number", "Declaration A Number",
        "Declaration Date", "Expected Receipt Date", "Pickup Address", "Is Dangerous",
        "Is Stackable", "Total Gross Weight (kg)", "Total Volumetric Weight (kg)",
        "Forwarder", "Created At",
        "Forwarder VRS",  # ✅ NEW LAST COLUMN (PART 1)
    ]
    ws.append(header_row_1)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    ws.append([
        imp.pk,
        imp.import_code,
        imp.vendor_name,
        str(imp.exporter_country) if imp.exporter_country else "",
        imp.incoterms or "",
        imp.currency_code or "",
        float(imp.goods_price) if imp.goods_price is not None else "",
        imp.shipping_method or "",
        imp.shipment_status or "",
        imp.tracking_no or "",
        imp.vendor_reference or "",
        imp.forwarder_reference or "",
        imp.declaration_c_number or "",
        imp.declaration_a_number or "",
        imp.declaration_date.isoformat() if imp.declaration_date else "",
        imp.expected_receipt_date.isoformat() if imp.expected_receipt_date else "",
        imp.pickup_address or "",
        "Yes" if imp.is_danger else "No",
        "Yes" if imp.is_stackable else "No",
        float(imp.total_gross_weight_kg) if imp.total_gross_weight_kg is not None else "",
        float(imp.total_volumetric_weight_kg) if imp.total_volumetric_weight_kg is not None else "",
        imp.forwarder.name if imp.forwarder else "",
        imp.created_at.strftime("%Y-%m-%d") if imp.created_at else "",
        forwarder_vrs_part1,  # ✅ VALUE FOR PART 1
    ])
    for cell in ws[2]:
        cell.border = THIN_BORDER

    ws.append([])

    # --------------------------
    # Sheet1 - Part 2 lines
    # --------------------------
    line_headers = [
        "Line ID", "Import ID", "Import Code", "Vendor", "Document No.", "Line No.",
        "Item No.", "Description", "Quantity", "Unit of Measure", "Unit Cost", "Line Amount",
        "Goods Currency", "Expected Receipt Date", "Delivery Date", "ATC Receipt Date",
        "Created At", "Shipment Status", "Tracking Number", "Vendor Reference",
        "Forwarder Reference", "Declaration C Number", "Declaration A Number",
        "Declaration Date", "Incoterms", "Total Import GW (kg)", "Total Import VW (kg)",
        "Line Gross Weight (kg)", "Line Volumetric Weight (kg)", "Vendor VRS",
        "Forwarder VRS",
    ]

    start_row_lines = ws.max_row + 1
    ws.append(line_headers)
    for cell in ws[start_row_lines]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    item_numbers = {l.item_no for l in lines if l.item_no}
    items = Item.objects.filter(number__in=item_numbers)
    item_map = {i.number: i for i in items}

    total_qty = sum((l.quantity or Decimal("0")) for l in lines) or Decimal("0")
    total_gw_import = imp.total_gross_weight_kg or Decimal("0")
    total_vw_import = imp.total_volumetric_weight_kg or Decimal("0")

    register_dt = imp.created_at

    forwarder_vrs_import = ""
    if register_dt and imp.declaration_date:
        forwarder_vrs_import = (register_dt.date() - imp.declaration_date).days

    for l in lines:
        qty = l.quantity or Decimal("0")
        item = item_map.get(l.item_no) if l.item_no else None

        line_gw = None
        line_vw = None

        if qty:
            if item and item.weight is not None:
                line_gw = item.weight * qty
            elif total_gw_import and total_qty:
                line_gw = (total_gw_import * qty) / total_qty

            if item and item.volumetric_weight is not None:
                line_vw = item.volumetric_weight * qty
            elif total_vw_import and total_qty:
                line_vw = (total_vw_import * qty) / total_qty

        vendor_vrs = ""
        if register_dt and l.delivery_date:
            vendor_vrs = (register_dt.date() - l.delivery_date).days

        ws.append([
            l.pk,
            imp.pk,
            imp.import_code,
            imp.vendor_name,
            l.document_no or "",
            l.line_no or "",
            l.item_no or "",
            l.description or "",
            float(l.quantity) if l.quantity is not None else "",
            l.unit_of_measure or "",
            float(l.unit_cost) if l.unit_cost is not None else "",
            float(l.line_amount) if l.line_amount is not None else "",
            imp.currency_code or "",
            l.expected_receipt_date.isoformat() if l.expected_receipt_date else "",
            l.delivery_date.isoformat() if l.delivery_date else "",
            imp.expected_receipt_date.isoformat() if imp.expected_receipt_date else "",
            imp.created_at.strftime("%Y-%m-%d") if imp.created_at else "",
            imp.shipment_status or "",
            imp.tracking_no or "",
            imp.vendor_reference or "",
            imp.forwarder_reference or "",
            imp.declaration_c_number or "",
            imp.declaration_a_number or "",
            imp.declaration_date.isoformat() if imp.declaration_date else "",
            imp.incoterms or "",
            float(total_gw_import) if total_gw_import else "",
            float(total_vw_import) if total_vw_import else "",
            float(line_gw) if line_gw is not None else "",
            float(line_vw) if line_vw is not None else "",
            vendor_vrs,
            forwarder_vrs_import,
        ])

        row_idx = ws.max_row
        fill_color = "FFFFFF" if row_idx % 2 == 0 else "F7F7F7"
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = THIN_BORDER

    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 2, 60)

    ws.freeze_panes = f"A{start_row_lines+1}"

    # --------------------------
    # Sheet2 packages
    # --------------------------
    ws2 = wb.create_sheet(title="Packages", index=1)
    pkg_headers = [
        "Package ID", "Import ID", "Import Code", "Package Type", "Length (cm)",
        "Width (cm)", "Height (cm)", "Gross Weight (kg)", "Unit System",
    ]
    ws2.append(pkg_headers)
    for cell in ws2[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    packages = ImportPackage.objects.filter(import_header=imp).order_by("pk")
    for pkg in packages:
        ws2.append([
            pkg.pk,
            imp.pk,
            imp.import_code,
            pkg.package_type or "",
            float(pkg.length_cm) if pkg.length_cm is not None else "",
            float(pkg.width_cm) if pkg.width_cm is not None else "",
            float(pkg.height_cm) if pkg.height_cm is not None else "",
            float(pkg.gross_weight_kg) if pkg.gross_weight_kg is not None else "",
            pkg.unit_system or "",
        ])

        row_idx = ws2.max_row
        fill_color = "FFFFFF" if row_idx % 2 == 0 else "F7F7F7"
        for cell in ws2[row_idx]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = THIN_BORDER

    for column_cells in ws2.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws2.column_dimensions[col_letter].width = min(length + 2, 60)

    ws2.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"import_{imp.import_code}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
