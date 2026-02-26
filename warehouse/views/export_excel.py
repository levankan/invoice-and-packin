from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from imports.models import Import, ImportLine

import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def download_excel(request):
    q = (request.GET.get("q") or "").strip()

    if q:
        # 🔥 Remove spaces just in case
        q = q.replace(" ", "")

        # 🔥 If starts with 20010 and does NOT contain "/", add it
        if q.startswith("20010") and "/" not in q:
            q = f"20010/{q[5:]}"

    results = Import.objects.none()
    lines = ImportLine.objects.none()

    if q:
        results = Import.objects.filter(
            Q(import_code__icontains=q) |
            Q(tracking_no__icontains=q) |
            Q(vendor_reference__icontains=q) |
            Q(forwarder_reference__icontains=q) |
            Q(declaration_c_number__icontains=q) |
            Q(declaration_a_number__icontains=q)
        ).distinct()

        if results.exists():
            lines = ImportLine.objects.filter(
                import_header__in=results
            ).select_related("import_header")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouse Export"

    headers = [
        "Import Code", "Document No", "Line No", "Item No", "Description",
        "Quantity", "UOM", "Tracking", "Declaration C", "Declaration A",
        "Vendor Ref", "Forwarder Ref"
    ]
    ws.append(headers)

    for l in lines:
        h = l.import_header
        ws.append([
            getattr(h, "import_code", ""),
            getattr(l, "document_no", ""),
            getattr(l, "line_no", ""),
            getattr(l, "item_no", ""),
            getattr(l, "description", ""),
            getattr(l, "quantity", ""),
            getattr(l, "unit_of_measure", ""),
            getattr(h, "tracking_no", ""),
            getattr(h, "declaration_c_number", ""),
            getattr(h, "declaration_a_number", ""),
            getattr(h, "vendor_reference", ""),
            getattr(h, "forwarder_reference", ""),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"warehouse_export_{q or 'all'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response