from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Q
from django.http import HttpResponse
from imports.models import Import, ImportLine

import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def dashboard(request):
    query = None
    results = None
    lines = None

    if request.method == "POST":
        query = request.POST.get("tracking_number", "").strip()

        if query:
            results = Import.objects.filter(
                Q(import_code__icontains=query) |
                Q(tracking_no__icontains=query) |
                Q(vendor_reference__icontains=query) |
                Q(forwarder_reference__icontains=query) |
                Q(declaration_c_number__icontains=query) |
                Q(declaration_a_number__icontains=query)
            ).distinct()

            if results.exists():
                lines = ImportLine.objects.filter(import_header__in=results)

    return render(request, "warehouse/dashboard.html", {
        "query": query,
        "results": results,
        "lines": lines,
    })


@login_required
def download_excel(request):
    q = (request.GET.get("q") or "").strip()

    # If no query provided, return empty file (or you can redirect back)
    results = Import.objects.none()
    lines = ImportLine.objects.none()

    if q:
        results = Import.objects.filter(
            Q(import_code__icontains=q) |
            Q(tracking_no__icontains=q) |
            Q(vendor_reference__icontains=q) |
            Q(forwarder_reference__icontains=q) |
            Q(declaration_c_number__icontains=q) |
            Q(declaration_a_number__icontains=q)
        ).distinct()

        if results.exists():
            lines = ImportLine.objects.filter(import_header__in=results).select_related("import_header")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouse Export"

    headers = [
        "Import Code", "Document No", "Line No", "Item No", "Description",
        "Quantity", "UOM", "Tracking", "Vendor Ref", "Forwarder Ref",
        "Declaration C", "Declaration A"
    ]
    ws.append(headers)

    # Fill rows (one row per ImportLine + its Import header)
    for l in lines:
        h = l.import_header  # Import header

        ws.append([
            getattr(h, "import_code", ""),
            getattr(l, "document_no", ""),
            getattr(l, "line_no", ""),
            getattr(l, "item_no", ""),
            getattr(l, "description", ""),
            getattr(l, "quantity", ""),
            getattr(l, "unit_of_measure", ""),
            getattr(h, "tracking_no", ""),
            getattr(h, "vendor_reference", ""),
            getattr(h, "forwarder_reference", ""),
            getattr(h, "declaration_c_number", ""),
            getattr(h, "declaration_a_number", ""),
        ])

    # Optional: auto width (simple)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Return as file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"warehouse_export_{q or 'all'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response